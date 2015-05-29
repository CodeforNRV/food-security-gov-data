from openpyxl import load_workbook

xlsx_name = '2014.xlsx'

wb = load_workbook(xlsx_name)

def proc_sheet(sheet_name, data_cols, r_start, r_end):
    sheet = wb.get_sheet_by_name(sheet_name)
    data = []
    for r in range(r_start, r_end+1):
        data.append(
            {field:sheet.cell(row=r, column=c).value for field,c in data_cols.iteritems()}
        )
        
    return data

data = {}

#Locality Designations
sheet_name = 'Locality Name'
data_cols ={
    "fips": 3,
    "name": 4,
    "region": 6
}
r_start = 9
r_end = 128

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)
    
#CSA LASER
sheet_name = 'CSA LASER'
data_cols = {
    "fips": 1,
    "CSA_STATE": 5,
    "CSA_LOCAL": 6
}
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#IT Support
sheet_name = 'IT Support'
data_cols = {
    "fips": 1,
    "level": 5
}
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)
    
#Agency Level
sheet_name = 'Agency Level'
data_cols = {
    "fips": 1,
    "level": 5
}
r_start = 4
r_end = 123

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#HR Policy
sheet_name = 'HR Policy'
data_cols = {
    "fips": 1,
    "status": 4
}
r_start = 4
r_end = 123

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Board Type
sheet_name = 'Board Type'
data_cols = {
    "fips": 1,
    "type": 4,
    "advisory_details": 5
}
r_start = 4
r_end = 123

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#2013 Population
sheet_name = '2013 Population'
data_cols = {
    "fips": 1,
    "child.white": 12,
    "child.black": 13,
    "child.asian": 14,
    "child.native": 15,
    "child.other": 16,
    "adult.white": 18,
    "adult.black": 19,
    "adult.asian": 20,
    "adult.native": 21,
    "adult.other": 22,
    "elderly.white": 24,
    "elderly.black": 25,
    "elderly.asian": 26,
    "elderly.native": 27,
    "elderly.other": 28,
    "hispanic.child": 38,
    "hispanic.adult": 39,
    "hispanic.elderly": 40
}
r_start = 6
r_end = 125

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Poverty_2013
sheet_name = 'Poverty_2013'
data_cols = {
    "fips": 1,
    "all.saipe_base": 4,
    "all.poverty_count": 5,
    "0-17.saipe_base": 7,
    "0-17.poverty_count": 8,
    "5-17.saipe_base": 10,
    "5-17.poverty_count": 11
}
r_start = 6
r_end = 125

temp = proc_sheet(sheet_name, data_cols, r_start, r_end)

data[sheet_name] = []
for datum in temp:
    data[sheet_name].append(
        {
            "fips": datum["fips"],
            "0-5.saipe_base": datum["0-17.saipe_base"] - datum["5-17.saipe_base"],
            "0-5.poverty_count": datum["0-17.poverty_count"] - datum["5-17.poverty_count"],
            "5-17.saipe_base": datum["5-17.saipe_base"],
            "5-17.poverty_count": datum["5-17.poverty_count"],
            "17+.saipe_base": datum["all.saipe_base"] - datum["0-17.saipe_base"],
            "17+.poverty_count": datum["all.poverty_count"] - datum["0-17.poverty_count"]
        }
    )
    
#Poverty_All ages
sheet_name = 'Poverty_All ages'
years = range(2000, 2014)
i_col = range(3, 17)
data_cols = { year: c for year,c in zip(years,i_col) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Poverty_Child
sheet_name = 'Poverty_Child'
years = range(2000, 2014)
i_col = range(18, 32)
data_cols = { year: c for year,c in zip(years,i_col) }
r_start = 6
r_end = 125

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Income to Poverty Ratio
sheet_name = "Income to Poverty Ratio"
offsets = {
    "<125%":0,
    ">=125%": 1
}
groups = {
    "0-5.": 3,
    "6-11.": 6,
    "12-17.": 9,
    "18-24.": 12,
    "25-34.": 15,
    "35-44.": 18,
    "45-54.": 21,
    "55-64.": 24,
    "65-74.": 27,
    "75=<.": 30
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 7
r_end = 126

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Unemployment
sheet_name = "Unemployment"

offsets = {
    "laborForce": 0,
    "numberUnemployed": 1
}
years = range(2000, 2014)
i_col = range(19, 46, 2)
groups = { str(year) + ".": c for year, c in zip(years, i_col) }
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
        
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Non-marital births
sheet_name = "Non-marital births"
offsets = {
    "white":0,
    "black": 1,
    "other": 2
}
groups = {
    "totalLive.": 5,
    "nonmaritalLive.": 9,
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 7
r_end = 126

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Teen births
sheet_name = "Teen births"
offsets = {
    "white":0,
    "black": 1,
    "other": 2
}
groups = {
    "teenBirths.": 5,
    "teenPopulation.": 9,
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 7
r_end = 126

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#NM Births_1998-2012
sheet_name = "NM Births_1998-2012"
years = range(1998, 2013)
offsets = { str(year) + ".": c for c, year in enumerate(years) }
groups = {
    "totalLive.": 4,
    "nonmaritalLive.": 19,
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 6
r_end = 125

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Teen Births_1998-2012
sheet_name = "Teen Births_1998-2012"
years = range(1998, 2013)
offsets = { str(year) + ".": c for c, year in enumerate(years) }
groups = {
    "teenBirths.": 4,
    "teenPopulation.": 19,
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 6
r_end = 125

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Children in Single Parent Homes
sheet_name = "Children in Single Parent Homes"
offsets = {
    "age<18": 0,
    "age<6": 11
}
groups = {
    "childrenWithMaried.": 5,
    "childrenWithFather.": 7,
    "childrenWithMother.": 8
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#SNAP Clients by SFY
sheet_name = 'SNAP Clients by SFY'
years = range(2005, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#TNAF Clients by SFY
sheet_name = 'TNAF Clients by SFY'
years = range(2005, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Midicaid Clients by SFY
sheet_name = 'Midicaid Clients by SFY'
years = range(2009, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#EA Clients by SFY
sheet_name = 'EA Clients by SFY'
years = range(2013, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Benefit Clients by SFY
sheet_name = 'Benefit Clients by SFY'
years = range(2009, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#ChildCareClients
sheet_name = 'ChildCareClients'
data_cols = {"2014": 4}
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#TNAF Cases by SFY
sheet_name = 'TNAF Cases by SFY'
years = range(2010, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#SNAP Cases by SFY
sheet_name = 'SNAP Cases by SFY'
years = range(2010, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Medicaid Cases by SFY
sheet_name = 'Medicaid Cases by SFY'
years = range(2010, 2015)
col_offset = 4
data_cols = { year: c + col_offset for c, year in enumerate(years) }
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#EA Cases by SFY
sheet_name = 'EA Cases by SFY'
offsets = {
    "fuel": 0,
    "cooling": 2,
    "crisis": 4
}
groups = {
    "2013.": 4,
    "2014.": 5
}
data_cols = {}
for group_label, group_col in groups.iteritems():
    for offset_label, offset_col in offsets.iteritems():
        data_cols[group_label+offset_label] = group_col+offset_col
r_start = 6
r_end = 125

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)

#Child Care Cases
sheet_name = 'Child Care Cases'
data_cols = {"2014": 4}
r_start = 5
r_end = 124

data[sheet_name] = proc_sheet(sheet_name, data_cols, r_start, r_end)
